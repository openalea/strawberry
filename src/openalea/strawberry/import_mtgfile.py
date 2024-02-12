from __future__ import absolute_import
from __future__ import print_function

import pandas as pd
import numpy as np
from openpyxl import load_workbook

from openalea.plantgl.all import*
from openalea.mtg import io
from openalea.mtg import algo
from openalea.lpy import *

import openalea.strawberry
from openalea.mtg import MTG, algo
from openalea.strawberry.data import data_directory


def name(f):
    """return base name without extension

    :param f: the file path
    :type f: string
    :return: basename
    :rtype: string
    """    
    return f.basename().splitext()[0]


def import_mtgfile(filename):
    """Import a MTG file from genotype name, in sharedata repo

    :param filename: genotype = name of the file
    :type filename: string
    :return: a MTG loaded from the file
    :rtype: MTG
    """    
    filenames = filename
    files = data_directory.glob('*.mtg')
    mtg_path = dict((name(f), f) for f in files)
    mtgfile = dict((k,f) for k,f in mtg_path.items() if k in filenames)
    if len(filenames) == 1:
        g = MTG(mtgfile[filenames[0]])
        return g
    else:
        metaMTG= MTG()
        for i in mtgfile:
            metaMTG = algo.union(metaMTG, MTG(mtgfile[i]))
        return metaMTG   

################
def read_file(file,sheet_name,convert=None):
    """read csv file in share data repo
    :param file: filename of the csv
    :type file: string
    :param sheet_name name of sheet csv file
    :type sheet_name: string
    :param convert (optional): colname conversion if need to change colname
    :type convert: dict
    :return: a panda dataframe load from file
    :rtype: panda.DataFram
    """
    xls= pd.ExcelFile(file)
    df=pd.read_excel(xls,sheet_name=sheet_name)
    
    return df

def filter_none(d):
    """ filter fonction for nan value
    :param d: dict of dict of mtg properties
    :type d: dict of dict contains for a dict of MTG properties name as key and for each key a dict with vids in key and property values  
    :return: a dict of dict of MTG properties without None values
    :rtype: dict of dict containing MTG properties
    """
    if isinstance(d, dict):
        return {k: filter_none(v) for k, v in d.items() if v!='None'}
    else:
        return d

def topology(df, first_property):
    """ Extract topology from csv file and transform it in string.
    This function contain a string_conversion test to check if have not error in the conversion to string
    :param df: data containing topological information
    :type df: pandas.DataFrame
    :param first_property: colname of the first properties associate to topology information
    :type first_property: string
    :return: string contain topological information
    :rtype: string
    """
    # select topological part of file
    start_prop=df.columns.get_loc(first_property)
    df = df.loc[:,list(df.columns[:start_prop])]
    
    # convert in string
    array = df.fillna(-1).to_numpy() #replace na value by -1 and convert to numpy array
    
    # convert array to string
    row_index = list(range(0,array.shape[0]))
    column_index = list(range(0,array.shape[1]))
    column_start = 0
    string=[]
    
    for row in row_index:
        for column in column_index:
            if array[row,column]!=-1:
                break
        if column == column_start:
            string.append(array[row,column])
        elif column < column_start:
            string.extend([']',array[row,column]])
        else:
            string.extend(['[',array[row,column]])
        
        column_start = column
      
    string = "".join(map(str,string))
    
    def test_string_convertion(string):
        """ test the converion into string counting the number of open and close brackets
        if the number of bracket == 0 conversion is ok
        elif the number of bracket is >0 means that we have more of open bracket than closed bracket and vice versa and print error of conversion
        
        :param string: string containing the transformation of csv into string
        :type string: string
        :return: nothing if ok or Error of conversion if error
        :rtype: nothing or error message
        """
        
        count=0
        for x in string:
            # print(x)
            if x == "[":
                count+=1
            elif x == "]":
                count-=1

        return count
    
    if test_string_convertion(string)==0:
        return string
    else:
        print("Error in convertion dataframe to string")
        

def add_properties(g, df, first_property):
    """add properties to MTG from csv
    
    :param g: MTG
    :type g: MTG object
    :param df: data frame containing properties
    :type df: pandas.DataFrame
    :param first_property: colname of the first property
    :type first_property: string
    
    :return: an MTG with his properties
    :rtype: MTG object
    """
   
    start_prop=df.columns.get_loc(first_property)
    df = df.loc[:,list(df.columns[start_prop:])]
    df=df.where(pd.notnull(df), None)
    df.index = np.arange(1, len(df) + 1)
    df=df.fillna('None')
    property_dict=df.to_dict()

    property_dict=filter_none(property_dict)

    for key,value in property_dict.items():
            g.properties()[key]=value


    return g

def add_axis_scale(g):
    """add axis scale in mtg from csv
    
    :param g: an MTG
    :type g: MTG object
    :return: an MTG with axis scale added
    :rtype: MTG object 
    """
    vids=g.vertices(scale=3)

    for vid in vids:
        pid = g.parent(vid)
        
        if g.edge_type(vid)=="+":
            cpx=g.add_child_and_complex(pid,vid)[-1]
            p_cpx = g.complex(pid)
            
            g.add_child(p_cpx, cpx, label="A", edge_type="+")
            
    g.reindex()
    return g

def strawberry_reader_csv(file, first_property='experimental_name',symbole_at_scale = dict(P=1,T=2, F=3, f=3, b=3, HT=3, bt=3, ht=3,s=3)):
    ''' Main function to import MTG from csv file
    
    :param file: filename of csv
    :type file: string
    :param first_property: colnames of the first property
    :type first_property: string
    :param symbole_at_scale: dict of label symbole and scale
    :type symbole_at_scale: dict
    :return: list of individuals mtg
    :rtype: a list of mtg by csv
    
    '''
    
    workbook = load_workbook(file)
    sheets= workbook.sheetnames
    
    mtgs = list()
    
    for sheet in sheets:
        df = read_file(file=file, sheet_name=sheet)  
        
        string= topology(df,first_property=first_property)
        scene = Scene()
        l = LsysContext()
        l.makeCurrent()
       
        for module in symbole_at_scale:
            l.declare(module)
           
        axialtree = AxialTree(string)
        g = io.axialtree2mtg(tree=axialtree, scale=symbole_at_scale, scene=scene)
        l.done()
        
        g= add_properties(g, df,first_property=first_property)
        g.property("Stade")
        g= add_axis_scale(g)
        g.properties()["order"] = algo.orders(g)
        mtgs.append(g)
        
        
    return mtgs

def import_mtg_from_csv(files,first_property,symbol_at_scale=dict(P=1,T=2, F=3, f=3, b=3, HT=3, bt=3, ht=3,s=3)):
    ''' import mtg from multiple csv file
    :param files: list of set of csv filenames
    :type files: list 
    :param first_property: name of the first property
    :type: string
    :param symbole_at_scale: dict of label symbole and scale
    :type: dict
    :return: dict of list of mtg by file
    :rtype: dict
    '''
    mtgs={}
    for file in files:
        name= str(file).split('\\')[-1].split(".")[0]
        print(name)
        mtg= strawberry_reader_csv(file)
        mtgs[name]=mtg
    return mtgs

def union(g):
    """ Union of all mtg in one
    :param g: MTGs
    :type g: list of dict of MTG
    :return: one MTG which are union of all mtg
    :rtype: MTG
    """
    if type(g) is list:
        mtg_list= g
    elif type(g) is dict:
        mtg_list = [mtg for sub_mtg in g.values() for mtg in sub_mtg]

    lg= algo.union(mtg_list[0],mtg_list[1])
    for mtg in mtg_list[2:]:
        lg = algo.union(lg,mtg)

    return lg
###############

def plant_number_by_varieties(g):
    """Print plant number by varieties in a mtg

    :param g: MTG
    :type g: MTG
    """    
    genotype = set(g.property("Genotype").values())

    for geno in genotype:
        no_plants= list(g.property("Genotype").values()).count(geno)
        print(geno, ":", no_plants, "plants")